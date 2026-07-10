"""Corrected-aware export wrapper.

Builds an export from a batch using corrected values (overlay), reuses the
existing workbook_from_rows() to keep main-sheet output byte-equivalent when
no corrections exist, then appends an 'Audit Changes' sheet and records a
BatchExportEvent. Status is moved to 'exported'.
"""
from __future__ import annotations
from copy import copy
from datetime import datetime
from io import BytesIO
import logging

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db import models as M
from app.db.review_models import InvoiceRowFieldAudit, BatchExportEvent, InvoiceFieldCandidate
from app.services import correction_service as cs
from app.services.candidate_outcomes import label_batch_candidates
from app.services.description_summary import summarise_total_invoice_description
from app.services.exporter import workbook_from_rows
from app.services.supplier_pattern_learning import promote_supplier_patterns_for_batch
from app.utils.storage import batch_export_folder

logger = logging.getLogger(__name__)


def _build_di_candidate_summary_map(db: Session, batch_id, scan_run_id=None) -> dict[int, str]:
    criteria = [
        InvoiceFieldCandidate.batch_id == batch_id,
        InvoiceFieldCandidate.source_type.in_(["azure_di", "azure_di_structured"]),
    ]
    if scan_run_id is not None:
        criteria.append(InvoiceFieldCandidate.scan_run_id == scan_run_id)
    rows = db.execute(
        select(InvoiceFieldCandidate).where(*criteria).order_by(
            InvoiceFieldCandidate.row_id.asc(),
            InvoiceFieldCandidate.field_name.asc(),
            InvoiceFieldCandidate.selected.desc(),
            InvoiceFieldCandidate.applied.desc(),
            InvoiceFieldCandidate.confidence.desc().nullslast(),
            InvoiceFieldCandidate.created_at.desc(),
        )
    ).scalars().all()
    grouped: dict[int, dict[str, InvoiceFieldCandidate]] = {}
    for cand in rows:
        field_map = grouped.setdefault(cand.row_id, {})
        if cand.field_name not in field_map:
            field_map[cand.field_name] = cand
    out: dict[int, str] = {}
    for row_id, field_map in grouped.items():
        parts: list[str] = []
        for field_name in ("supplier_name", "invoice_number", "invoice_date", "net_amount", "vat_amount", "total_amount", "currency"):
            cand = field_map.get(field_name)
            if not cand or not cand.candidate_value:
                continue
            conf = ""
            if cand.confidence is not None:
                try:
                    conf = f" ({float(cand.confidence):.0%})"
                except Exception:
                    conf = ""
            evidence = (cand.evidence or "").strip()
            if evidence:
                evidence = evidence[:120]
                parts.append(f"{field_name}={cand.candidate_value}{conf} [{evidence}]")
            else:
                parts.append(f"{field_name}={cand.candidate_value}{conf}")
        if parts:
            out[row_id] = " | ".join(parts)[:4000]
    return out


def build_corrected_rows(db: Session, batch: M.InvoiceBatch) -> list[dict]:
    scan_run_id = getattr(batch, "current_scan_run_id", None)
    row_query = select(M.InvoiceRow).where(
        M.InvoiceRow.batch_id == batch.id,
        M.InvoiceRow.row_status == M.INVOICE_ROW_STATUS_ACTIVE,
    )
    if scan_run_id is not None:
        row_query = row_query.where(M.InvoiceRow.scan_run_id == scan_run_id)
    rows = db.execute(
        row_query.order_by(M.InvoiceRow.source_file_id, M.InvoiceRow.page_no, M.InvoiceRow.id)
    ).scalars().all()
    cmap = cs.load_correction_map(db, batch.id)
    di_summary_by_row = _build_di_candidate_summary_map(db, batch.id, scan_run_id)
    out = []
    for r in rows:
        c = cmap.get(r.id)
        d = {col: getattr(r, col) for col in r.__table__.columns.keys()}
        if c is not None:
            for f in (
                "supplier_name", "supplier_posting_account", "nominal_account_code",
                "invoice_number", "invoice_date", "description",
                "net_amount", "vat_amount", "total_amount", "currency", "tax_code",
            ):
                v = getattr(c, f, None)
                if v is not None:
                    d[f] = v
        if (getattr(batch, "scan_mode", None) or "summary").lower() == "summary":
            d["description"] = summarise_total_invoice_description(d.get("description"), d.get("line_items_raw"))
        d["di_candidate_summary"] = di_summary_by_row.get(r.id)
        out.append(d)
    return out


def export_batch_corrected(
    db: Session,
    *,
    batch: M.InvoiceBatch,
    user: M.User,
    template_sheet=None,
    nominal_account_map: dict[str, str] | None = None,
    batch_metadata: dict | None = None,
) -> BytesIO:
    """Render the workbook with corrected values, append audit sheet, log event."""
    rows = build_corrected_rows(db, batch)
    base_buf: BytesIO = workbook_from_rows(
        rows,
        batch_metadata=batch_metadata,
        nominal_account_map=nominal_account_map,
        template_sheet=template_sheet,
    )

    # Reopen workbook to append audit sheet
    base_buf.seek(0)
    wb = load_workbook(base_buf)
    ws = wb.create_sheet("Audit Changes")
    headers = ["Batch ID", "Export Version", "Exported At", "Row ID", "Field",
               "Original / Old", "New / Current", "Action", "User", "Note", "Changed At"]
    ws.append(headers)

    next_version = (getattr(batch, "current_export_version", 0) or 0) + 1
    now = datetime.utcnow().isoformat()

    audits = db.execute(
        select(InvoiceRowFieldAudit)
        .where(InvoiceRowFieldAudit.batch_id == batch.id)
        .order_by(InvoiceRowFieldAudit.created_at)
    ).scalars().all()
    for a in audits:
        ws.append([
            str(batch.id), next_version, now, a.row_id, a.field_name,
            a.old_value or "", a.new_value or "", a.action,
            a.username or (str(a.user_id) if a.user_id else ""),
            a.note or "", a.created_at.isoformat(),
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    export_folder = batch_export_folder(batch.id)
    export_filename = f"batch_{batch.id}_v{next_version}.xlsx"
    export_path = export_folder / export_filename
    export_bytes = out.getvalue()
    export_path.write_bytes(export_bytes)
    out.seek(0)

    # Log export event + flip status + bump version
    logger.info(
        "export completed batch=%s version=%d rows=%d audit_rows=%d bytes=%d",
        batch.id, next_version, len(rows), len(audits), len(export_bytes),
    )

    labelled_candidates = label_batch_candidates(db, batch=batch, user=user, outcome_source="export")
    logger.info("candidate outcomes labelled batch=%s labels=%d source=export", batch.id, labelled_candidates)
    promoted_patterns = promote_supplier_patterns_for_batch(db, batch=batch, user=user, outcome_source="export")
    logger.info("supplier pattern learning promoted batch=%s patterns=%d source=export", batch.id, promoted_patterns)

    ev = BatchExportEvent(
        batch_id=batch.id, scan_run_id=getattr(batch, "current_scan_run_id", None), export_version=next_version,
        exported_by=user.id, exported_at=datetime.utcnow(),
        file_path=str(export_path), file_bytes=export_bytes, storage_backend="database+local", row_count=len(rows),
    )
    db.add(ev)
    batch.current_export_version = next_version
    batch.exported_at = datetime.utcnow()
    batch.exported_by = user.id
    batch.status = "exported"
    db.add(InvoiceRowFieldAudit(
        batch_id=batch.id, scan_run_id=getattr(batch, "current_scan_run_id", None), row_id=0, field_name="__export__",
        old_value=None, new_value=f"v{next_version}", action="export",
        user_id=user.id,
    ))
    return out
