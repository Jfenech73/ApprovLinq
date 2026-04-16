"""
Tests for the Export Template module.

Tests the render service, transform logic, workbook integration, and assignment
precedence — all as pure unit tests (no database or network required).
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import types
from datetime import date, datetime
from io import BytesIO

import pandas as pd
import pytest

from app.services.template_render_service import (
    AVAILABLE_FIELDS,
    COLUMN_TYPES,
    apply_transform,
    render_template_sheet,
    _sanitize_sheet_name,
)
from app.services.exporter import workbook_from_rows


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_column(
    col_id,
    heading,
    col_type,
    source_field=None,
    static_value=None,
    transform_rule=None,
    is_active=True,
    col_order=0,
    notes=None,
):
    col = types.SimpleNamespace(
        id=col_id,
        column_heading=heading,
        column_type=col_type,
        source_field=source_field,
        static_value=static_value,
        transform_rule=transform_rule,
        is_active=is_active,
        column_order=col_order,
        notes=notes,
    )
    return col


def _make_template(name, columns):
    tpl = types.SimpleNamespace(name=name, columns=columns)
    return tpl


def _make_invoice_row(**kwargs):
    defaults = {
        "supplier_name": "Test Supplier Ltd",
        "invoice_number": "INV-001",
        "invoice_date": date(2024, 3, 15),
        "net_amount": 850.00,
        "vat_amount": 195.50,
        "total_amount": 1045.50,
        "currency": "EUR",
        "nominal_account_code": "5001",
        "validation_status": "ok",
        "review_required": False,
    }
    defaults.update(kwargs)
    return defaults


# ── apply_transform ───────────────────────────────────────────────────────────

class TestApplyTransform:
    def test_no_rule_returns_unchanged(self):
        assert apply_transform("hello", None) == "hello"

    def test_uppercase(self):
        assert apply_transform("acme ltd", "uppercase") == "ACME LTD"

    def test_uppercase_on_none_returns_none(self):
        assert apply_transform(None, "uppercase") is None

    def test_lowercase(self):
        assert apply_transform("ACME LTD", "lowercase") == "acme ltd"

    def test_number_format_coerces_string(self):
        assert apply_transform("1234.56", "number_format") == 1234.56

    def test_number_format_non_numeric_unchanged(self):
        assert apply_transform("N/A", "number_format") == "N/A"

    def test_date_format_applies_strftime(self):
        d = date(2024, 3, 15)
        assert apply_transform(d, "date_format:%d/%m/%Y") == "15/03/2024"

    def test_date_format_on_non_date_unchanged(self):
        assert apply_transform("not a date", "date_format:%d/%m/%Y") == "not a date"

    def test_default_fallback_used_when_none(self):
        assert apply_transform(None, "default:N/A") == "N/A"

    def test_default_fallback_used_when_empty_string(self):
        assert apply_transform("", "default:MISSING") == "MISSING"

    def test_default_fallback_not_used_when_value_present(self):
        assert apply_transform("Acme", "default:N/A") == "Acme"

    def test_unknown_rule_returns_value_unchanged(self):
        assert apply_transform("hello", "unknown_rule_xyz") == "hello"


# ── _sanitize_sheet_name ──────────────────────────────────────────────────────

class TestSanitizeSheetName:
    def test_truncates_to_31_chars(self):
        name = "A" * 50
        result = _sanitize_sheet_name(name)
        assert len(result) <= 31

    def test_replaces_illegal_chars(self):
        result = _sanitize_sheet_name("Sheet:Name/With\\Illegal?Chars*[]!")
        for ch in r":\/?*[]":
            assert ch not in result

    def test_empty_string_gives_export(self):
        assert _sanitize_sheet_name("") == "Export"

    def test_short_name_unchanged(self):
        assert _sanitize_sheet_name("Sage 50") == "Sage 50"


# ── render_template_sheet ─────────────────────────────────────────────────────

class TestRenderTemplateSheet:
    def _basic_template(self):
        cols = [
            _make_column(1, "Supplier",      "mapped_field",  source_field="supplier_name", col_order=0),
            _make_column(2, "Invoice No",    "mapped_field",  source_field="invoice_number", col_order=1),
            _make_column(3, "Journal",       "static_text",   static_value="PURCHASES",     col_order=2),
            _make_column(4, "Reference 2",   "empty_column",                                col_order=3),
            _make_column(5, "Net Amount",    "mapped_field",  source_field="net_amount",    col_order=4),
        ]
        return _make_template("Sage 50 Import", cols)

    def test_returns_correct_sheet_name(self):
        tpl = self._basic_template()
        sheet_name, _ = render_template_sheet(tpl, [_make_invoice_row()])
        assert sheet_name == "Sage 50 Import"

    def test_mapped_field_resolved(self):
        tpl = self._basic_template()
        _, rows = render_template_sheet(tpl, [_make_invoice_row()])
        assert rows[0]["Supplier"] == "Test Supplier Ltd"
        assert rows[0]["Invoice No"] == "INV-001"

    def test_static_text_repeated(self):
        tpl = self._basic_template()
        _, rows = render_template_sheet(tpl, [_make_invoice_row(), _make_invoice_row()])
        assert rows[0]["Journal"] == "PURCHASES"
        assert rows[1]["Journal"] == "PURCHASES"

    def test_empty_column_is_none(self):
        tpl = self._basic_template()
        _, rows = render_template_sheet(tpl, [_make_invoice_row()])
        assert rows[0]["Reference 2"] is None

    def test_inactive_column_excluded(self):
        cols = [
            _make_column(1, "Supplier", "mapped_field", source_field="supplier_name", col_order=0, is_active=True),
            _make_column(2, "Hidden",   "mapped_field", source_field="invoice_number", col_order=1, is_active=False),
        ]
        tpl = _make_template("Test", cols)
        _, rows = render_template_sheet(tpl, [_make_invoice_row()])
        assert "Supplier" in rows[0]
        assert "Hidden" not in rows[0]

    def test_column_order_respected(self):
        cols = [
            _make_column(1, "Z-Last",  "static_text", static_value="z", col_order=99),
            _make_column(2, "A-First", "static_text", static_value="a", col_order=0),
        ]
        tpl = _make_template("Ordered", cols)
        _, rows = render_template_sheet(tpl, [_make_invoice_row()])
        keys = list(rows[0].keys())
        assert keys.index("A-First") < keys.index("Z-Last")

    def test_missing_mapped_field_returns_none(self):
        cols = [_make_column(1, "Missing", "mapped_field", source_field="nonexistent_field", col_order=0)]
        tpl = _make_template("Test", cols)
        _, rows = render_template_sheet(tpl, [_make_invoice_row()])
        assert rows[0]["Missing"] is None

    def test_bool_coerced_to_yes_no(self):
        cols = [_make_column(1, "Review?", "mapped_field", source_field="review_required", col_order=0)]
        tpl = _make_template("Test", cols)
        _, rows = render_template_sheet(tpl, [_make_invoice_row(review_required=False)])
        assert rows[0]["Review?"] == "No"

    def test_enrichment_merged_into_rows(self):
        cols = [_make_column(1, "Company", "mapped_field", source_field="company_name", col_order=0)]
        tpl = _make_template("Test", cols)
        _, rows = render_template_sheet(
            tpl,
            [_make_invoice_row()],
            enrichment={"company_name": "Acme Ltd"},
        )
        assert rows[0]["Company"] == "Acme Ltd"

    def test_transform_applied_to_mapped_field(self):
        cols = [_make_column(1, "Supplier UC", "mapped_field", source_field="supplier_name", transform_rule="uppercase", col_order=0)]
        tpl = _make_template("Test", cols)
        _, rows = render_template_sheet(tpl, [_make_invoice_row(supplier_name="Acme Ltd")])
        assert rows[0]["Supplier UC"] == "ACME LTD"

    def test_empty_rows_list_returns_empty(self):
        tpl = self._basic_template()
        _, rows = render_template_sheet(tpl, [])
        assert rows == []

    def test_multiple_rows(self):
        tpl = self._basic_template()
        invoice_rows = [
            _make_invoice_row(invoice_number="INV-001"),
            _make_invoice_row(invoice_number="INV-002"),
            _make_invoice_row(invoice_number="INV-003"),
        ]
        _, rows = render_template_sheet(tpl, invoice_rows)
        assert len(rows) == 3
        assert rows[2]["Invoice No"] == "INV-003"


# ── Derived / conditional value columns ───────────────────────────────────────

class TestDerivedAndConditional:
    def test_derived_value_uses_source_field_and_transform(self):
        cols = [_make_column(1, "STATUS", "derived_value", source_field="validation_status", transform_rule="uppercase", col_order=0)]
        tpl = _make_template("T", cols)
        _, rows = render_template_sheet(tpl, [_make_invoice_row(validation_status="ok")])
        assert rows[0]["STATUS"] == "OK"

    def test_conditional_value_with_default_fallback(self):
        cols = [_make_column(1, "Tax Code", "conditional_value", source_field="tax_code", transform_rule="default:T1", col_order=0)]
        tpl = _make_template("T", cols)
        _, rows = render_template_sheet(tpl, [_make_invoice_row(tax_code=None)])
        assert rows[0]["Tax Code"] == "T1"


# ── Workbook integration ──────────────────────────────────────────────────────

class TestWorkbookWithTemplateSheet:
    def _fake_row(self):
        row = types.SimpleNamespace()
        for col in [
            "source_filename", "page_no", "supplier_name", "supplier_posting_account",
            "nominal_account_code", "invoice_number", "invoice_date", "description",
            "line_items_raw", "net_amount", "vat_amount", "total_amount", "currency",
            "tax_code", "method_used", "confidence_score", "validation_status",
            "review_required", "review_priority", "review_reasons", "review_fields",
            "auto_approved", "page_quality_score", "classification_method",
            "supplier_match_method", "totals_reconciliation_status", "header_raw",
            "totals_raw", "page_text_raw", "id", "batch_id", "tenant_id",
            "company_id", "source_file_id", "created_at",
        ]:
            setattr(row, col, None)
        row.page_no = 1
        row.review_required = False
        row.auto_approved = False
        row.supplier_name = "Test Supplier"
        row.net_amount = 100.0
        row.vat_amount = 23.0
        row.total_amount = 123.0
        row.__table__ = types.SimpleNamespace(
            columns=types.SimpleNamespace(
                keys=lambda: [
                    "id", "batch_id", "tenant_id", "company_id", "source_file_id",
                    "source_filename", "page_no", "supplier_name", "supplier_posting_account",
                    "nominal_account_code", "invoice_number", "invoice_date", "description",
                    "line_items_raw", "net_amount", "vat_amount", "total_amount", "currency",
                    "tax_code", "method_used", "confidence_score", "validation_status",
                    "review_required", "review_priority", "review_reasons", "review_fields",
                    "auto_approved", "page_quality_score", "classification_method",
                    "supplier_match_method", "totals_reconciliation_status", "header_raw",
                    "totals_raw", "page_text_raw", "created_at",
                ]
            )
        )
        return row

    def test_workbook_without_template_returns_four_sheets(self):
        row = self._fake_row()
        buf = workbook_from_rows([row])
        buf.seek(0)
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        assert set(wb.sheetnames) == {"Invoices", "Needs Review", "Summary", "Evidence"}

    def test_workbook_with_template_adds_fifth_sheet(self):
        row = self._fake_row()
        tpl_df = pd.DataFrame([{"Supplier": "Test Supplier", "Journal": "PURCHASES", "Ref2": ""}])
        buf = workbook_from_rows([row], template_sheet=("Sage 50", tpl_df))
        buf.seek(0)
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        assert "Sage 50" in wb.sheetnames
        assert len(wb.sheetnames) == 5

    def test_workbook_template_sheet_has_correct_headers(self):
        row = self._fake_row()
        tpl_df = pd.DataFrame([{"Account Code": "5001", "Journal": "PURCHASES", "Amount": 123.0}])
        buf = workbook_from_rows([row], template_sheet=("Xero Import", tpl_df))
        buf.seek(0)
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        ws = wb["Xero Import"]
        headers = [cell.value for cell in ws[1]]
        assert headers == ["Account Code", "Journal", "Amount"]


# ── Available fields catalogue ────────────────────────────────────────────────

class TestAvailableFields:
    def test_catalogue_is_non_empty_list_of_strings(self):
        assert isinstance(AVAILABLE_FIELDS, list)
        assert len(AVAILABLE_FIELDS) > 0
        assert all(isinstance(f, str) for f in AVAILABLE_FIELDS)

    def test_common_invoice_fields_present(self):
        required = [
            "supplier_name", "invoice_number", "invoice_date",
            "net_amount", "vat_amount", "total_amount",
            "currency", "nominal_account_code", "validation_status",
        ]
        for f in required:
            assert f in AVAILABLE_FIELDS, f"Expected {f!r} in AVAILABLE_FIELDS"


# ── Column type catalogue ─────────────────────────────────────────────────────

class TestColumnTypes:
    def test_all_required_types_present(self):
        expected = {"mapped_field", "static_text", "empty_column", "derived_value", "conditional_value"}
        assert expected == COLUMN_TYPES


# ── Regression: export without template still works ───────────────────────────

class TestExportRegression:
    def test_no_template_workbook_identical_to_original(self):
        """Existing export must be completely unaffected when no template is assigned."""
        import types as t
        row = types.SimpleNamespace()
        for col in [
            "id", "batch_id", "tenant_id", "company_id", "source_file_id",
            "source_filename", "page_no", "supplier_name", "supplier_posting_account",
            "nominal_account_code", "invoice_number", "invoice_date", "description",
            "line_items_raw", "net_amount", "vat_amount", "total_amount", "currency",
            "tax_code", "method_used", "confidence_score", "validation_status",
            "review_required", "review_priority", "review_reasons", "review_fields",
            "auto_approved", "page_quality_score", "classification_method",
            "supplier_match_method", "totals_reconciliation_status", "header_raw",
            "totals_raw", "page_text_raw", "created_at",
        ]:
            setattr(row, col, None)
        row.page_no = 1
        row.review_required = False
        row.auto_approved = False
        row.__table__ = types.SimpleNamespace(
            columns=types.SimpleNamespace(
                keys=lambda: [
                    "id", "batch_id", "tenant_id", "company_id", "source_file_id",
                    "source_filename", "page_no", "supplier_name", "supplier_posting_account",
                    "nominal_account_code", "invoice_number", "invoice_date", "description",
                    "line_items_raw", "net_amount", "vat_amount", "total_amount", "currency",
                    "tax_code", "method_used", "confidence_score", "validation_status",
                    "review_required", "review_priority", "review_reasons", "review_fields",
                    "auto_approved", "page_quality_score", "classification_method",
                    "supplier_match_method", "totals_reconciliation_status", "header_raw",
                    "totals_raw", "page_text_raw", "created_at",
                ]
            )
        )
        buf = workbook_from_rows([row], template_sheet=None)
        assert isinstance(buf, BytesIO)
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        assert "Invoices" in wb.sheetnames
        assert "Sage 50" not in wb.sheetnames


# ── Regression: condition_rules absent on lightweight column object (Defect E) ─

class TestConditionalValueWithoutConditionRules:
    """Defect E: render_template_sheet must not raise AttributeError when the
    column object (e.g. a SimpleNamespace built in tests) lacks a
    condition_rules attribute. The fix uses getattr(col, 'condition_rules', None)."""

    def test_conditional_no_condition_rules_attr_falls_back_to_transform(self):
        """Column has NO condition_rules attribute at all — must fall back to transform_rule."""
        col = _make_column(
            1, "Tax Code", "conditional_value",
            source_field="tax_code",
            transform_rule="default:T1",
            col_order=0,
        )
        # _make_column produces a SimpleNamespace — condition_rules is absent by design
        assert not hasattr(col, "condition_rules")
        tpl = _make_template("T", [col])
        _, rows = render_template_sheet(tpl, [_make_invoice_row(tax_code=None)])
        # transform_rule default:T1 must apply
        assert rows[0]["Tax Code"] == "T1"

    def test_conditional_condition_rules_none_falls_back_to_transform(self):
        """Column has condition_rules=None explicitly — must fall back to transform."""
        col = _make_column(
            2, "Tax Code", "conditional_value",
            source_field="tax_code",
            transform_rule="default:T2",
            col_order=0,
        )
        col.condition_rules = None          # explicitly set to None
        tpl = _make_template("T", [col])
        _, rows = render_template_sheet(tpl, [_make_invoice_row(tax_code=None)])
        assert rows[0]["Tax Code"] == "T2"

    def test_conditional_condition_rules_empty_list_falls_back_to_transform(self):
        """Empty condition_rules list is falsy — must still fall back to transform."""
        col = _make_column(
            3, "Tax Code", "conditional_value",
            source_field="tax_code",
            transform_rule="default:T3",
            col_order=0,
        )
        col.condition_rules = []
        tpl = _make_template("T", [col])
        _, rows = render_template_sheet(tpl, [_make_invoice_row(tax_code=None)])
        assert rows[0]["Tax Code"] == "T3"

    def test_conditional_with_populated_condition_rules_evaluates_normally(self):
        """When condition_rules is a non-empty list, conditions are evaluated as before."""
        col = _make_column(
            4, "Flag", "conditional_value",
            source_field="currency",
            col_order=0,
        )
        col.condition_rules = [
            {"if_field": "currency", "operator": "eq", "value": "EUR", "output": "EURO"},
            {"output": "OTHER"},
        ]
        tpl = _make_template("T", [col])
        _, rows_eur = render_template_sheet(tpl, [_make_invoice_row(currency="EUR")])
        _, rows_gbp = render_template_sheet(tpl, [_make_invoice_row(currency="GBP")])
        assert rows_eur[0]["Flag"] == "EURO"
        assert rows_gbp[0]["Flag"] == "OTHER"

    def test_no_attribute_error_raised_without_condition_rules(self):
        """Explicit check: calling render_template_sheet with a SimpleNamespace
        column missing condition_rules must never raise AttributeError."""
        col = _make_column(
            5, "X", "conditional_value",
            source_field="invoice_number",
            col_order=0,
        )
        tpl = _make_template("T", [col])
        try:
            render_template_sheet(tpl, [_make_invoice_row()])
        except AttributeError as exc:
            pytest.fail(f"render_template_sheet raised AttributeError: {exc}")


# ── Regression: Authorization header binding (Defect D) ──────────────────────

class TestAuthorizationHeaderBinding:
    """Defect D: review.py endpoints used plain `authorization: str | None = None`
    which does NOT bind the HTTP Authorization header in FastAPI.

    These tests verify the fix by inspecting the function signatures directly —
    no running server required.
    """

    def test_current_user_flexible_authorization_uses_header(self):
        """current_user_flexible must declare authorization via Header(...)."""
        import inspect
        from fastapi import Header
        from app.routers.review import current_user_flexible

        sig = inspect.signature(current_user_flexible)
        param = sig.parameters.get("authorization")
        assert param is not None, "current_user_flexible must have an 'authorization' parameter"
        default = param.default
        # FastAPI Header parameters are wrapped in FieldInfo; check the type name
        assert "Header" in type(default).__name__ or hasattr(default, "alias"), (
            f"'authorization' parameter must use Header(...), got {type(default).__name__}"
        )

    def test_file_info_authorization_uses_header(self):
        """file_info must declare authorization via Header(...)."""
        import inspect
        from app.routers.review import file_info

        sig = inspect.signature(file_info)
        param = sig.parameters.get("authorization")
        assert param is not None, "file_info must have an 'authorization' parameter"
        default = param.default
        assert "Header" in type(default).__name__ or hasattr(default, "alias"), (
            f"file_info 'authorization' must use Header(...), got {type(default).__name__}"
        )

    def test_preview_authorization_uses_header(self):
        """preview must declare authorization via Header(...)."""
        import inspect
        from app.routers.review import preview

        sig = inspect.signature(preview)
        param = sig.parameters.get("authorization")
        assert param is not None, "preview must have an 'authorization' parameter"
        default = param.default
        assert "Header" in type(default).__name__ or hasattr(default, "alias"), (
            f"preview 'authorization' must use Header(...), got {type(default).__name__}"
        )
