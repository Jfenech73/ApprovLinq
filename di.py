#!/usr/bin/env python3
"""Dump raw Microsoft Document Intelligence fields, tables, and page lines from a PDF."""

import argparse
import csv
import json
import os
import sys
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request


# Fill these in, then run:
#   py di_pdf_dump.py
#
# Command-line arguments still work and override these values.
PDF_PATH_PLACEHOLDER = r"C:\path\to\invoice.pdf"
DI_ENDPOINT_PLACEHOLDER = r"https://YOUR-RESOURCE.cognitiveservices.azure.com"
DI_KEY_PLACEHOLDER = r"YOUR_DOCUMENT_INTELLIGENCE_KEY"

PDF_PATH = r"C:\Users\user\Documents\GitHub\Invoice Scanning\Files\Miriana1.pdf"
DI_ENDPOINT = r"https://approvlinq.cognitiveservices.azure.com"
DI_KEY = r"IVNrhyf4EmVNRrSJSVTIU5BzkhUtZT36mhG6fpyRzujTfHvMVCa2JQQJ99CCACPV0roXJ3w3AAALACOGm5hp"
DI_MODEL = "prebuilt-invoice"
DI_API_VERSION = "2024-11-30"
DI_PAGES = ""  # Examples: "1", "1-5", "1,3,7-9". Set to "" to read the full PDF.
PAGES_PER_BATCH = 1  # Used when DI_PAGES is empty. Sends pages 1-3, 4-6, 7-9, etc.
BATCH_DELAY_SECONDS = 20  # Pause between each DI batch to avoid rate/size pressure.
SPLIT_BEFORE_UPLOAD = True  # Requires: py -m pip install pypdf
JSON_OUT = r"raw_di_output.json"  # Set to "" if you do not want the full JSON saved.
ROW_OUT = r"di_rows.tsv"  # Excel-friendly row output. Set to "" to only print to screen.
EXCEL_OUT = r"di_output.xlsx"  # Real Excel output with preserved columns.


def parse_pages(pages_text):
    pages = []
    for part in pages_text.replace(" ", "").split(","):
        if not part:
            continue
        if "-" in part:
            start, end = part.split("-", 1)
            pages.extend(range(int(start), int(end) + 1))
        else:
            pages.append(int(part))
    return sorted(set(pages))


def get_pdf_page_count(pdf_path):
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise SystemExit(
            "pypdf is required for large PDF batching.\n"
            "Install it with: py -m pip install pypdf\n"
        ) from exc

    return len(PdfReader(pdf_path).pages)


def make_page_subset_pdf(pdf_path, pages_text):
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError as exc:
        raise SystemExit(
            "SPLIT_BEFORE_UPLOAD is True but pypdf is not installed.\n"
            "Install it with: py -m pip install pypdf\n"
            "Or set SPLIT_BEFORE_UPLOAD = False and use DI_PAGES only."
        ) from exc

    selected_pages = parse_pages(pages_text)
    reader = PdfReader(pdf_path)
    writer = PdfWriter()
    for page_number in selected_pages:
        if page_number < 1 or page_number > len(reader.pages):
            raise SystemExit(f"Page {page_number} is outside the PDF page range 1-{len(reader.pages)}.")
        writer.add_page(reader.pages[page_number - 1])

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    temp_file.close()
    with open(temp_file.name, "wb") as f:
        writer.write(f)
    return temp_file.name


def page_chunks(pdf_path, pages_text, pages_per_batch):
    if pages_text:
        pages = parse_pages(pages_text)
    else:
        pages = list(range(1, get_pdf_page_count(pdf_path) + 1))

    if pages_per_batch <= 0:
        raise SystemExit("PAGES_PER_BATCH must be 1 or higher.")

    return [pages[i:i + pages_per_batch] for i in range(0, len(pages), pages_per_batch)]


def format_pages(pages):
    if not pages:
        return ""
    ranges = []
    start = previous = pages[0]
    for page in pages[1:]:
        if page == previous + 1:
            previous = page
            continue
        ranges.append(f"{start}-{previous}" if start != previous else str(start))
        start = previous = page
    ranges.append(f"{start}-{previous}" if start != previous else str(start))
    return ",".join(ranges)


def request_json(url, key, method="GET", body=None, content_type="application/json"):
    req = urllib.request.Request(url, data=body, method=method)
    req.add_header("Ocp-Apim-Subscription-Key", key)
    if body is not None:
        req.add_header("Content-Type", content_type)
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = resp.read()
            return resp, json.loads(data.decode("utf-8")) if data else {}
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise SystemExit(f"HTTP {exc.code} from Document Intelligence:\n{detail}") from exc


def analyze_pdf(endpoint, key, pdf_path, model, api_version, pages):
    endpoint = endpoint.rstrip("/")
    url = f"{endpoint}/documentintelligence/documentModels/{model}:analyze?api-version={api_version}"
    if pages and not SPLIT_BEFORE_UPLOAD:
        url += f"&pages={urllib.parse.quote(pages)}"
    with open(pdf_path, "rb") as f:
        resp, _ = request_json(url, key, method="POST", body=f.read(), content_type="application/pdf")
    operation_url = resp.headers.get("operation-location")
    if not operation_url:
        raise SystemExit("Document Intelligence did not return an operation-location header.")

    while True:
        _, result = request_json(operation_url, key)
        status = result.get("status")
        if status in {"succeeded", "failed", "canceled"}:
            if status != "succeeded":
                raise SystemExit(json.dumps(result, indent=2, ensure_ascii=False))
            return result.get("analyzeResult", result)
        time.sleep(1.5)


def field_value(field):
    for key in ("content", "valueString", "valueDate", "valueTime", "valuePhoneNumber", "valueNumber",
                "valueInteger", "valueCurrency", "valueAddress", "valueSelectionMark"):
        if key in field:
            return field[key]
    if "valueArray" in field:
        return f"[array: {len(field['valueArray'])}]"
    if "valueObject" in field:
        return f"[object: {len(field['valueObject'])} keys]"
    return ""


def printable_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else str(value)


def build_row_output(result):
    rows = []
    documents = result.get("documents", [])
    if not documents:
        return [["No documents returned by Document Intelligence."]]

    for d_i, doc in enumerate(documents, 1):
        fields = doc.get("fields", {})
        header_fields = {k: v for k, v in fields.items() if k != "Items"}
        header_names = ["DocumentNumber", "DocType", "DocumentConfidence"] + sorted(header_fields)
        header_values = [d_i, doc.get("docType", ""), doc.get("confidence", "")]
        header_values.extend(printable_value(field_value(header_fields[name])) for name in sorted(header_fields))

        rows.append([f"HEADER DOCUMENT {d_i}"])
        rows.append(header_names)
        rows.append(header_values)
        rows.append([])

        items = fields.get("Items", {}).get("valueArray", [])
        detail_names = set()
        detail_values = []
        for item in items:
            values = item.get("valueObject", {})
            detail_names.update(values.keys())
            detail_values.append(values)

        detail_columns = ["DocumentNumber", "LineNumber", "LineConfidence"] + sorted(detail_names)
        rows.append([f"DETAILS DOCUMENT {d_i}"])
        rows.append(detail_columns)
        for line_number, values in enumerate(detail_values, 1):
            row = [d_i, line_number, items[line_number - 1].get("confidence", "")]
            row.extend(printable_value(field_value(values.get(name, {}))) for name in sorted(detail_names))
            rows.append(row)
        rows.append([])

    return rows


def collect_excel_rows(result, batch_pages):
    header_rows = []
    detail_rows = []
    documents = result.get("documents", [])
    batch_text = format_pages(batch_pages)

    for d_i, doc in enumerate(documents, 1):
        fields = doc.get("fields", {})
        header = {
            "BatchPages": batch_text,
            "DocumentInBatch": d_i,
            "DocType": doc.get("docType", ""),
            "DocumentConfidence": doc.get("confidence", ""),
        }
        for name, field in fields.items():
            if name != "Items":
                header[name] = printable_value(field_value(field))
        header_rows.append(header)

        items = fields.get("Items", {}).get("valueArray", [])
        for line_number, item in enumerate(items, 1):
            row = {
                "BatchPages": batch_text,
                "DocumentInBatch": d_i,
                "LineNumber": line_number,
                "LineConfidence": item.get("confidence", ""),
            }
            for name, field in item.get("valueObject", {}).items():
                row[name] = printable_value(field_value(field))
            detail_rows.append(row)

    return header_rows, detail_rows


def collect_table_rows(result, batch_pages):
    rows = []
    batch_text = format_pages(batch_pages)
    for table_index, table in enumerate(result.get("tables", []), 1):
        for cell in table.get("cells", []):
            rows.append({
                "BatchPages": batch_text,
                "TableNumber": table_index,
                "RowIndex": cell.get("rowIndex", ""),
                "ColumnIndex": cell.get("columnIndex", ""),
                "Kind": cell.get("kind", ""),
                "Content": cell.get("content", ""),
                "Confidence": cell.get("confidence", ""),
            })
    return rows


def collect_page_line_rows(result, batch_pages):
    rows = []
    first_page = batch_pages[0] if batch_pages else 0
    batch_text = format_pages(batch_pages)
    for page in result.get("pages", []):
        output_page = first_page + page.get("pageNumber", 1) - 1
        for line_number, line in enumerate(page.get("lines", []), 1):
            rows.append({
                "BatchPages": batch_text,
                "PageNumber": output_page,
                "LineNumber": line_number,
                "Content": line.get("content", ""),
            })
    return rows


def write_excel_output(path, header_rows, detail_rows, table_rows, page_line_rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required for .xlsx output.\n"
            "Install it with: py -m pip install openpyxl"
        ) from exc

    def write_sheet(wb, title, rows):
        ws = wb.create_sheet(title)
        columns = []
        for row in rows:
            for key in row.keys():
                if key not in columns:
                    columns.append(key)
        if not columns:
            ws.append(["No rows"])
            return
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(column, "") for column in columns])
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Headers", header_rows)
    write_sheet(wb, "Details", detail_rows)
    write_sheet(wb, "Tables", table_rows)
    write_sheet(wb, "PageLines", page_line_rows)
    wb.save(path)
    print(f"Saved Excel output to {path}")


def write_row_output(rows, path=None):
    writer = csv.writer(sys.stdout, delimiter="\t", lineterminator="\n")
    writer.writerows(rows)

    if path:
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            file_writer = csv.writer(f, delimiter="\t", lineterminator="\n")
            file_writer.writerows(rows)
        print(f"\nSaved row output to {path}")


def print_tables(result):
    print("\n=== TABLES ===")
    for t_i, table in enumerate(result.get("tables", []), 1):
        rows = table.get("rowCount", 0)
        cols = table.get("columnCount", 0)
        print(f"\nTable {t_i}: {rows} rows x {cols} cols")
        grid = [["" for _ in range(cols)] for _ in range(rows)]
        for cell in table.get("cells", []):
            r, c = cell.get("rowIndex", 0), cell.get("columnIndex", 0)
            if r < rows and c < cols:
                grid[r][c] = cell.get("content", "").replace("\n", " ")
        for r, row in enumerate(grid):
            print(f"  row {r}: " + " | ".join(row))


def print_page_lines(result):
    print("\n=== PAGE LINES ===")
    for page in result.get("pages", []):
        print(f"\nPage {page.get('pageNumber', '?')}:")
        for line in page.get("lines", []):
            print(f"  {line.get('content', '')}")


def main():
    parser = argparse.ArgumentParser(description="Dump Microsoft Document Intelligence output for one PDF.")
    parser.add_argument("pdf", nargs="?", default=PDF_PATH, help="Path to the PDF file to analyze")
    parser.add_argument("--endpoint", default=os.getenv("AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT") or DI_ENDPOINT)
    parser.add_argument("--key", default=os.getenv("AZURE_DOCUMENT_INTELLIGENCE_KEY") or DI_KEY)
    parser.add_argument("--model", default=DI_MODEL, help="DI model id, e.g. prebuilt-invoice")
    parser.add_argument("--api-version", default=DI_API_VERSION)
    parser.add_argument("--pages", default=DI_PAGES, help='Page range, e.g. "1-5" or "1,3,7-9". Empty means all pages.')
    parser.add_argument("--pages-per-batch", type=int, default=PAGES_PER_BATCH, help="Number of PDF pages to send per DI request.")
    parser.add_argument("--batch-delay", type=float, default=BATCH_DELAY_SECONDS, help="Seconds to pause between DI batches.")
    parser.add_argument("--no-split", action="store_true", help="Do not create a smaller page-subset PDF before upload.")
    parser.add_argument("--json-out", default=JSON_OUT, help="Optional path to save the full raw analyzeResult JSON")
    parser.add_argument("--row-out", default=ROW_OUT, help="Optional path to save Excel-friendly row output as TSV")
    parser.add_argument("--excel-out", default=EXCEL_OUT, help="Optional path to save real Excel output as XLSX")
    args = parser.parse_args()

    if not args.endpoint or args.endpoint == DI_ENDPOINT_PLACEHOLDER:
        raise SystemExit("Set DI_ENDPOINT at the top of this script, pass --endpoint, or set AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT.")
    if not args.key or args.key == DI_KEY_PLACEHOLDER:
        raise SystemExit("Set DI_KEY at the top of this script, pass --key, or set AZURE_DOCUMENT_INTELLIGENCE_KEY.")
    if not args.pdf or args.pdf == PDF_PATH_PLACEHOLDER:
        raise SystemExit("Set PDF_PATH at the top of this script or pass a PDF path as the first argument.")
    if not os.path.isfile(args.pdf):
        raise SystemExit(f"PDF not found: {args.pdf}")

    all_results = []
    header_rows = []
    detail_rows = []
    table_rows = []
    page_line_rows = []
    batches = page_chunks(args.pdf, args.pages, args.pages_per_batch)
    for batch_index, batch_pages in enumerate(batches, 1):
        batch_text = format_pages(batch_pages)
        pdf_to_send = args.pdf
        temp_pdf = None
        if SPLIT_BEFORE_UPLOAD and not args.no_split:
            temp_pdf = make_page_subset_pdf(args.pdf, batch_text)
            pdf_to_send = temp_pdf
            print(f"Analyzing PDF pages {batch_text} using temporary file {temp_pdf}")
        else:
            print(f"Analyzing PDF pages {batch_text}")

        result = analyze_pdf(args.endpoint, args.key, pdf_to_send, args.model, args.api_version, batch_text)
        all_results.append({"pages": batch_text, "result": result})
        batch_headers, batch_details = collect_excel_rows(result, batch_pages)
        header_rows.extend(batch_headers)
        detail_rows.extend(batch_details)
        table_rows.extend(collect_table_rows(result, batch_pages))
        page_line_rows.extend(collect_page_line_rows(result, batch_pages))

        if temp_pdf:
            os.remove(temp_pdf)

        if batch_index < len(batches) and args.batch_delay > 0:
            print(f"Waiting {args.batch_delay:g} seconds before next batch...")
            time.sleep(args.batch_delay)

    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump(all_results, f, indent=2, ensure_ascii=False)
        print(f"Saved full raw analyzeResult JSON to {args.json_out}")

    rows = []
    for batch in all_results:
        rows.append([f"PAGES {batch['pages']}"])
        rows.extend(build_row_output(batch["result"]))
    write_row_output(rows, args.row_out)
    if args.excel_out:
        write_excel_output(args.excel_out, header_rows, detail_rows, table_rows, page_line_rows)


if __name__ == "__main__":
    main()
